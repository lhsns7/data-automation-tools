#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""xlsx.py — 서식 있는 엑셀 출력 (2026-08)
납품용 엑셀의 기본기를 한 곳에 모은 것. CSV로 주면 고객이 열 때마다 서식·인코딩으로 고생한다.

  from xlsx import write_workbook
  write_workbook('결과.xlsx', {'데이터': (헤더리스트, 행리스트)}, summary={'수집 시각': ..., '총 건수': ...})

들어가는 것: 헤더 고정(틀 고정)·자동 필터·열 너비 자동·숫자/날짜 서식·머리행 스타일·요약 시트·원자적 저장(임시파일→교체).
한글 CSV의 엑셀 깨짐(cp949) 문제를 원천 회피한다.
"""
import os, re, tempfile, datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEAD_FILL = PatternFill('solid', fgColor='1F4E79')
HEAD_FONT = Font(color='FFFFFF', bold=True, size=10)
BORDER = Border(bottom=Side(style='thin', color='D9D9D9'))
RE_INT = re.compile(r'^-?\d{1,15}$')
RE_FLOAT = re.compile(r'^-?\d+\.\d+$')
RE_DATE = re.compile(r'^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$')
RE_DTM = re.compile(r'^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})[ T](\d{1,2}):(\d{2})')


def coerce(v):
    """문자열을 엑셀이 아는 타입으로. 실패하면 원본 문자열 그대로(억지 변환 금지)."""
    if v is None or isinstance(v, (int, float, dt.date, dt.datetime)):
        return v
    s = str(v).strip()
    if not s:
        return None
    if RE_INT.match(s):
        # ★2026-08-29 검거: 앞자리가 0인 코드(종목코드 005930, 고유번호 00554352)를 숫자로 바꾸면 0이 사라진다.
        #   엑셀 납품에서 가장 흔한 클레임이라 문자로 유지한다. 아주 긴 숫자도 지수표기를 피해 문자로 둔다.
        if (s[0] == '0' and len(s) > 1) or (s[:2] == '-0' and len(s) > 2):
            return s
        n = int(s)
        return n if abs(n) < 10 ** 15 else s
    if RE_FLOAT.match(s):
        return float(s)
    m = RE_DTM.match(s)
    if m:
        try: return dt.datetime(*[int(x) for x in m.groups()])
        except ValueError: return s
    m = RE_DATE.match(s)
    if m:
        try: return dt.date(*[int(x) for x in m.groups()])
        except ValueError: return s
    return s


def _fmt(v):
    if isinstance(v, dt.datetime): return 'yyyy-mm-dd hh:mm'
    if isinstance(v, dt.date): return 'yyyy-mm-dd'
    if isinstance(v, int): return '#,##0'
    if isinstance(v, float): return '#,##0.00##'
    return None


def _sheet(ws, header, rows, freeze=True, autofilter=True, max_width=52):
    ws.append(list(header))
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c); cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
    widths = [len(str(h)) + 2 for h in header]
    for r in rows:
        vals = [coerce(x) for x in r]
        ws.append(vals)
        for i, v in enumerate(vals[:len(widths)]):
            f = _fmt(v)
            cell = ws.cell(row=ws.max_row, column=i + 1)
            if f: cell.number_format = f
            cell.border = BORDER
            ln = len(f'{v:,}') if isinstance(v, int) else len(str(v)) if v is not None else 0
            if isinstance(v, (dt.date, dt.datetime)): ln = 17
            if ln > widths[i]: widths[i] = min(ln + 2, max_width)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(w, 8)
    ws.row_dimensions[1].height = 22
    if freeze: ws.freeze_panes = 'A2'
    if autofilter and rows: ws.auto_filter.ref = f'A1:{get_column_letter(len(header))}{ws.max_row}'


def write_workbook(path, sheets, summary=None, summary_title='요약'):
    """sheets = {시트명: (헤더, 행들)}. summary = {항목: 값} → 맨 앞 시트로 삽입.
    원자적 저장: 임시파일에 쓰고 교체 → 중간에 죽어도 기존 파일이 깨지지 않는다."""
    wb = Workbook(); wb.remove(wb.active)
    if summary:
        ws = wb.create_sheet(summary_title)
        _sheet(ws, ['항목', '값'], [[k, v] for k, v in summary.items()], autofilter=False)
        ws.column_dimensions['A'].width = 26; ws.column_dimensions['B'].width = 60
    total = 0
    for name, (header, rows) in sheets.items():
        _sheet(wb.create_sheet(name[:31]), header, rows); total += len(rows)
    d = os.path.dirname(os.path.abspath(path)) or '.'
    os.makedirs(d, exist_ok=True)
    fd, tmp = tempfile.mkstemp(prefix='.tmp_', suffix='.xlsx', dir=d); os.close(fd)
    try:
        wb.save(tmp); os.replace(tmp, path)
    except Exception:
        try: os.remove(tmp)
        except OSError: pass
        raise
    return {'path': path, 'sheets': len(sheets) + (1 if summary else 0), 'rows': total,
            'size_kb': round(os.path.getsize(path) / 1024, 1)}

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""branch_merge.py — 지점·부서 엑셀 취합기 (양식 이탈 흡수·중복 제출 방지·누락 명단)

본사가 지점들에게 같은 양식 파일을 받아 합치는 흔한 업무의 자동화. 현실은 양식대로 오지 않는다:
컬럼명을 바꾸고, 순서를 섞고, 위에 제목을 얹고, 중간에 합계 행을 넣고, 같은 지점이 두 번 보낸다.

설계 (조용한 오염 방지가 본체):
  - 양식 흡수: 컬럼명 동의어 매핑('판매수량'='수량') · 순서 무관 · 여분 컬럼은 버리되 보고
  - 헤더 자동 탐지: 상단 제목·빈 행을 건너뛰고 표준 컬럼이 맞는 행을 헤더로
  - 합계 행 제외: '합계/총계' 행을 데이터로 합산하지 않고 제외 보고(이중 계상 방지)
  - ★중복 제출: 같은 지점 파일 여러 개 → **최신본만 채택 + 보고** (조용한 이중 합산 = 최악 사고)
  - 필수 컬럼 누락 파일 = 격리(합산 0·사유 보고) · ★등록부 대비 미제출 지점 명단
  - ★합계 항등: 통합 총액 = 채택 지점별 원본 합의 합 (자체 대사)

검증(--make-demo): 지점 8곳 등록 · 파일 7개 제출 시뮬(정상 2·양식 변형·잡행+합계행·필수 누락·
  중복 제출 2본) → ①취합 전수 ②변형 흡수 ③잡행/합계행 처리 ④최신본 채택·이중 합산 0
  ⑤격리·누락 명단 ⑥합계 항등+재현성.
※ 표준 양식·동의어·지점 등록부 = 설정값(고객 양식에 1회 맞춤).
"""
import os, sys, csv, shutil, datetime as dt

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(HERE, '..', '..', 'core'))
from xlsx import write_workbook
import openpyxl

SCHEMA = ['날짜', '상품', '수량', '금액']                     # 표준 양식(설정값)
SYNONYM = {'일자': '날짜', '품명': '상품', '판매수량': '수량', '판매금액': '금액'}
BRANCHES = ['강남', '서초', '판교', '분당', '일산', '수원', '부산', '대구']   # 등록부(설정값)
TOTAL_WORDS = ('합계', '총계', '소계')


def norm_col(c):
    c = str(c or '').strip()
    return SYNONYM.get(c, c)


def read_rows(path):
    """xlsx/csv → 원시 행 리스트(list of list)"""
    if path.lower().endswith('.csv'):
        return [row for row in csv.reader(open(path, encoding='utf-8-sig'))]
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    return [[c.value for c in row] for row in ws.iter_rows()]


def parse_file(path):
    """→ dict(rows=[표준화 dict], extra_cols, skipped_top, total_rows_removed, error)"""
    raw = read_rows(path)
    header_i, colmap = None, None
    for i, row in enumerate(raw[:10]):                     # 헤더 탐지: 표준 컬럼 3개+ 매칭되는 첫 행
        normed = [norm_col(c) for c in row]
        hit = sum(1 for c in normed if c in SCHEMA)
        if hit >= 3:
            header_i = i
            colmap = {j: c for j, c in enumerate(normed) if c}
            break
    if header_i is None:
        return dict(rows=[], extra_cols=[], skipped_top=0, totals=0,
                    error='헤더를 찾지 못함(표준 컬럼 불일치)')
    normed_cols = set(colmap.values())
    missing = [c for c in SCHEMA if c not in normed_cols]
    if missing:
        return dict(rows=[], extra_cols=[], skipped_top=header_i, totals=0,
                    error=f'필수 컬럼 누락: {", ".join(missing)}')
    extra = sorted(normed_cols - set(SCHEMA))
    rows, totals = [], 0
    for row in raw[header_i + 1:]:
        cells = {colmap.get(j): v for j, v in enumerate(row) if colmap.get(j)}
        first = str(row[0] or '').strip() if row else ''
        if not any(v not in (None, '') for v in (row or [])):
            continue                                        # 빈 행
        if any(w in first for w in TOTAL_WORDS):
            totals += 1                                     # 합계 행 = 데이터 아님(이중 계상 방지)
            continue
        try:
            rows.append({'날짜': str(cells['날짜']).strip()[:10],
                         '상품': str(cells['상품']).strip(),
                         '수량': int(str(cells['수량']).replace(',', '')),
                         '금액': int(str(cells['금액']).replace(',', ''))})
        except (ValueError, KeyError, TypeError):
            rows.append(None)                               # 불량 행 표식
    bad = rows.count(None)
    rows = [r for r in rows if r]
    return dict(rows=rows, extra_cols=extra, skipped_top=header_i, totals=totals,
                bad_rows=bad, error='')


def branch_of(fname):
    """파일명 규칙 '지점_*.csv|xlsx' → 지점명(설정값 규칙)"""
    base = os.path.basename(fname)
    return base.split('_', 1)[0] if '_' in base else os.path.splitext(base)[0]


def merge(inbox, branches=BRANCHES):
    """제출함 폴더 취합 → dict(merged, per_branch, dup_report, quarantined, missing, notes)"""
    files = sorted(os.path.join(inbox, f) for f in os.listdir(inbox)
                   if f.lower().endswith(('.csv', '.xlsx')))
    by_branch = {}
    for p in files:
        by_branch.setdefault(branch_of(p), []).append(p)
    merged, per_branch, dup_report, quarantined, notes = [], {}, [], [], []
    for br, ps in sorted(by_branch.items()):
        ps_sorted = sorted(ps, key=os.path.getmtime)
        chosen = ps_sorted[-1]                              # ★최신본 채택
        if len(ps) > 1:
            dup_report.append((br, len(ps), os.path.basename(chosen),
                               ', '.join(os.path.basename(x) for x in ps_sorted[:-1])))
        r = parse_file(chosen)
        if r['error']:
            quarantined.append((br, os.path.basename(chosen), r['error']))
            continue
        for row in r['rows']:
            merged.append(dict(row, 지점=br))
        per_branch[br] = dict(n=len(r['rows']), amt=sum(x['금액'] for x in r['rows']),
                              file=os.path.basename(chosen))
        if r['extra_cols']:
            notes.append(f'{br}: 여분 컬럼 {", ".join(r["extra_cols"])} (무시·보고)')
        if r['skipped_top']:
            notes.append(f'{br}: 상단 잡행 {r["skipped_top"]}행 건너뜀')
        if r['totals']:
            notes.append(f'{br}: 합계 행 {r["totals"]}개 제외(이중 계상 방지)')
        if r.get('bad_rows'):
            notes.append(f'{br}: 불량 행 {r["bad_rows"]}개 격리')
    submitted = set(by_branch)
    missing = [b for b in branches if b not in submitted]
    return dict(merged=merged, per_branch=per_branch, dup_report=dup_report,
                quarantined=quarantined, missing=missing, notes=notes)


def write_out(res, out):
    mrows = [[r['지점'], r['날짜'], r['상품'], r['수량'], r['금액']] for r in res['merged']]
    brows = [[b, v['n'], v['amt'], v['file']] for b, v in sorted(res['per_branch'].items())]
    sheets = {'통합': (['지점', '날짜', '상품', '수량', '금액'], mrows),
              '지점별 요약': (['지점', '행수', '금액 합', '채택 파일'], brows)}
    audit = ([['중복 제출', f'{b}: {n}개 → {c} 채택 (미채택: {o})'] for b, n, c, o in res['dup_report']]
             + [['격리', f'{b} ({f}): {e}'] for b, f, e in res['quarantined']]
             + [['미제출', ', '.join(res['missing']) or '없음']]
             + [['참고', n] for n in res['notes']])
    sheets['검수 리포트'] = (['구분', '내용'], audit)
    total = sum(r['금액'] for r in res['merged'])
    per_sum = sum(v['amt'] for v in res['per_branch'].values())
    return write_workbook(out, sheets, summary={
        '생성': dt.datetime.now().strftime('%Y-%m-%d %H:%M'),
        '취합': f"{len(res['per_branch'])}지점 {len(res['merged'])}행 · 격리 {len(res['quarantined'])} · 미제출 {len(res['missing'])}",
        '★합계 항등': f'통합 {total:,} = 지점합 {per_sum:,} → {"PASS" if total == per_sum else "★FAIL"}',
        '규칙': '동의어 매핑·순서 무관 · 합계행 제외 · 중복 제출=최신본 · 필수 누락=격리'})


# ── 검증 데모 ───────────────────────────────────────────────────────
def _csv(path, header, rows):
    with open(path, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)


def _xlsx(path, rows_of_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows_of_rows:
        ws.append(row)
    wb.save(path)


def make_demo(inbox):
    if os.path.isdir(inbox):
        shutil.rmtree(inbox)
    os.makedirs(inbox)
    mk = lambda br, i: [f'2026-09-0{(i%3)+1}', ['비타민C', '유산균', '텀블러'][i % 3], i % 4 + 1, (i % 4 + 1) * 12000]
    planted = {}
    # 정상 2곳(서초 csv · 수원 xlsx)
    rows = [mk('서초', i) for i in range(6)]
    _csv(os.path.join(inbox, '서초_판매.csv'), SCHEMA, rows)
    planted['서초'] = (6, sum(r[3] for r in rows))
    rows = [mk('수원', i) for i in range(5)]
    _xlsx(os.path.join(inbox, '수원_판매.xlsx'), [SCHEMA] + rows)
    planted['수원'] = (5, sum(r[3] for r in rows))
    # 양식 변형(판교): 컬럼명 변경+순서 셔플+여분 '비고'
    rows = [mk('판교', i) for i in range(7)]
    shuffled = [['품명', '판매금액', '비고', '일자', '판매수량']] + \
               [[r[1], r[3], f'메모{i}', r[0], r[2]] for i, r in enumerate(rows)]
    _xlsx(os.path.join(inbox, '판교_판매.xlsx'), shuffled)
    planted['판교'] = (7, sum(r[3] for r in rows))
    # 잡행+합계행(분당): 제목 2행 + 중간 합계 1행
    rows = [mk('분당', i) for i in range(4)]
    body = [['9월 지점 판매 보고'], [None], SCHEMA] + rows[:2] + \
           [['합계', '', '', sum(r[3] for r in rows[:2])]] + rows[2:]
    _xlsx(os.path.join(inbox, '분당_판매.xlsx'), body)
    planted['분당'] = (4, sum(r[3] for r in rows))
    # 필수 누락(일산): '금액' 없음 → 격리
    _csv(os.path.join(inbox, '일산_판매.csv'), ['날짜', '상품', '수량'],
         [['2026-09-01', '비타민C', 2]])
    # ★중복 제출(강남): 구본(5행) → 신본(6행, 마커 상품 포함) — mtime으로 신구 구분
    old_rows = [mk('강남', i) for i in range(5)]
    _csv(os.path.join(inbox, '강남_판매_v1.csv'), SCHEMA, old_rows)
    new_rows = [mk('강남', i) for i in range(5)] + [['2026-09-03', '신본마커상품', 1, 99000]]
    _csv(os.path.join(inbox, '강남_판매_v2.csv'), SCHEMA, new_rows)
    t = dt.datetime(2026, 9, 3, 8, 0).timestamp()
    os.utime(os.path.join(inbox, '강남_판매_v1.csv'), (t - 3600, t - 3600))
    os.utime(os.path.join(inbox, '강남_판매_v2.csv'), (t, t))
    planted['강남'] = (6, sum(r[3] for r in new_rows))
    return planted


def main_demo():
    inbox = os.path.join(HERE, 'demo_inbox')
    planted = make_demo(inbox)
    res = merge(inbox)
    res2 = merge(inbox)

    # ① 취합 전수: 채택 5지점, 지점별 행수·금액 = 심은 값
    ok1 = (set(res['per_branch']) == {'서초', '수원', '판교', '분당', '강남'}
           and all(res['per_branch'][b]['n'] == n and res['per_branch'][b]['amt'] == amt
                   for b, (n, amt) in planted.items()))
    # ② 양식 변형 흡수(판교): 값까지 정확 + 여분 컬럼 보고
    pangyo = [r for r in res['merged'] if r['지점'] == '판교']
    ok2 = (len(pangyo) == 7 and sum(r['금액'] for r in pangyo) == planted['판교'][1]
           and any('판교' in n and '비고' in n for n in res['notes']))
    # ③ 잡행/합계행(분당): 데이터 4행 · 합계행 1 제외 · 잡행 스킵 보고
    ok3 = (res['per_branch']['분당']['n'] == 4
           and any('분당' in n and '합계 행 1개' in n for n in res['notes'])
           and any('분당' in n and '잡행' in n for n in res['notes']))
    # ④ 중복 제출(강남): 신본 채택(마커 존재) · 이중 합산 0(6행) · 보고
    gangnam = [r for r in res['merged'] if r['지점'] == '강남']
    ok4 = (len(gangnam) == 6 and any(r['상품'] == '신본마커상품' for r in gangnam)
           and len(res['dup_report']) == 1 and res['dup_report'][0][0] == '강남'
           and 'v2' in res['dup_report'][0][2])
    # ⑤ 격리(일산 — 금액 누락) + 미제출 명단(부산·대구)
    ok5 = (len(res['quarantined']) == 1 and res['quarantined'][0][0] == '일산'
           and '금액' in res['quarantined'][0][2] and res['missing'] == ['부산', '대구'])
    # ⑥ 합계 항등 + 재현성
    total = sum(r['금액'] for r in res['merged'])
    ok6 = (total == sum(amt for _, amt in planted.values())
           and res['merged'] == res2['merged'])

    out = os.path.join(HERE, '통합_데모.xlsx')
    info = write_out(res, out)

    L = [f'# 지점 취합 검증 리포트 ({dt.datetime.now():%Y-%m-%d %H:%M})',
         '- 데모 = 등록 8지점 · 제출 파일 7개(csv+xlsx 혼합) — 양식 변형·잡행·합계행·필수 누락·**중복 제출** 전부 심음',
         '', '| 검증 | 결과 |', '|---|---|',
         f'| ① 취합 전수(5지점 행수·금액 = 심은 값) | {"PASS" if ok1 else "★FAIL"} |',
         f'| ② 양식 변형 흡수(컬럼명·순서 셔플·여분 컬럼 보고) | {"PASS" if ok2 else "★FAIL"} |',
         f'| ③ 상단 잡행 스킵 + 합계 행 제외(이중 계상 방지) | {"PASS" if ok3 else "★FAIL"} |',
         f'| ④ ★중복 제출 = 최신본만 채택(마커 확인·6행)·보고 | {"PASS" if ok4 else "★FAIL"} |',
         f'| ⑤ 필수 누락 격리(일산) + 미제출 명단(부산·대구) | {"PASS" if ok5 else "★FAIL"} |',
         f'| ⑥ 합계 항등(통합 {total:,} = 지점합) + 재현성 | {"PASS" if ok6 else "★FAIL"} |',
         f'| 산출 | {os.path.basename(out)} ({info["sheets"]}시트) |',
         '', '- ※ 표준 양식·동의어·지점 등록부 = 설정값(고객 양식 1회 맞춤). 격리·중복·미제출은 묵살하지 않고 검수 시트에 보고.']
    rep = os.path.join(HERE, 'branch_merge_verify.md')
    open(rep, 'w', encoding='utf-8').write('\n'.join(L))
    print('\n'.join(L))
    return ok1 and ok2 and ok3 and ok4 and ok5 and ok6


if __name__ == '__main__':
    sys.stdout.reconfigure(encoding='utf-8')
    ok = main_demo()
    sys.exit(0 if ok else 1)
